import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from datetime import date, datetime, timedelta
from io import BytesIO

def get_housekeeping_dates(arrival, departure):
    dates = []
    current = arrival + timedelta(days=3)
    while current < departure:
        dates.append(current)
        current += timedelta(days=3)
    return dates

def should_get_housekeeping_today(arrival, departure, today):
    return today.date() in [d.date() for d in get_housekeeping_dates(arrival, departure)]


def apply_excel_formatting(ws, guest_data_dict, today):
    """Formats the new workbook and populates data."""
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    vertical_only_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style=None), bottom=Side(style=None)
    )
    center_aligned_text = Alignment(horizontal='center', vertical='center')
    bold_font = Font(bold=True)
    x_font = Font(bold=True)  # Bold 'X' for HSK column

    # Set column widths
    col_widths = {
        'A': 7, 'B': 21, 'C': 10, 'D': 10, 'E': 10,'F': 3, 
        'G': 7, 'H': 21, 'I': 10, 'J': 10, 'K': 10
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Set a consistent height for data rows (Rows 2 to 32)
    for row_num in range(2, 34):
        ws.row_dimensions[row_num].height = 14.5

    # Loop through the cells once to apply border and alignment
    # Skip row 2 (the empty row) when applying borders
    for row in range(3, 34):  # Start from row 3 to skip the empty row
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            # Apply special border for column 6 (the empty separator column)
            if col == 6:
                cell.border = vertical_only_border
            else:
                cell.border = thin_border
            cell.alignment = center_aligned_text

    # Add Date and Headers
    bold_font = Font(bold=True)
    
    # Apply date with bold formatting
    ws['B1'] = 'GUEST LIST DATE:'
    ws['B1'].alignment = Alignment(horizontal='right')
    
    ws['C1'] = date.today().strftime('%Y-%m-%d')
    ws['C1'].font = bold_font
    
    headers = ['ROOM', 'GUEST NAME', 'ARRIVE', 'DEPART', 'HSK', '',
               'ROOM', 'GUEST NAME', 'ARRIVE', 'DEPART', 'HSK']
    start_col, start_row = 1, 3  # Changed from 2 to 3 to add empty line
    for i, header in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + i, value=header)
        cell.font = bold_font

    # Room numbers for column A (Left side)
    room_numbers_col_1 = [105, 106, 107, 108, 109, 110, 111, 112, 114, 115,
                          201, 202, 203, 204, 205, 206, 207, 208, 209, 210,
                          211, 212, 214, 215, 216, 217, 218, 219, 220, 221]
    # Room numbers for column H (Right side)
    room_numbers_col_8 = [222, 223, 224, 225, 226,
                          301, 302, 303, 304, 305, 306, 307, 308, 309, 310,
                          311, 312, 314, 315, 316, 317, 318, 319, 320,
                          321, 322, 323, 324, 325, 326]

    start_row_rooms = 4

    # Populate room numbers for both columns
    for i, room_number in enumerate(room_numbers_col_1):
        cell = ws.cell(row=start_row_rooms + i, column=1, value=room_number)
        cell.font = bold_font
    for i, room_number in enumerate(room_numbers_col_8):
        cell = ws.cell(row=start_row_rooms + i, column=7, value=room_number)
        cell.font = bold_font

    # Populate guest data
    for room, guest_info in guest_data_dict.items():
        arrival_str = guest_info['Arrival_Date'].strftime('%Y-%m-%d')
        depart_str = guest_info['Departure_Date'].strftime('%Y-%m-%d')

        if room in room_numbers_col_1:
            row_idx = start_row_rooms + room_numbers_col_1.index(room)
            ws.cell(row=row_idx, column=2, value=guest_info['Guest_Name'])
            ws.cell(row=row_idx, column=3, value=arrival_str)
            ws.cell(row=row_idx, column=4, value=depart_str)
            if should_get_housekeeping_today(guest_info['Arrival_Date'], guest_info['Departure_Date'], today):
                cell = ws.cell(row=row_idx, column=5, value='X')
                cell.font = x_font

        elif room in room_numbers_col_8:
            row_idx = start_row_rooms + room_numbers_col_8.index(room)
            ws.cell(row=row_idx, column=8, value=guest_info['Guest_Name'])
            ws.cell(row=row_idx, column=9, value=arrival_str)
            ws.cell(row=row_idx, column=10, value=depart_str)
            if should_get_housekeeping_today(guest_info['Arrival_Date'], guest_info['Departure_Date'], today):
                cell = ws.cell(row=row_idx, column=11, value='X')
                cell.font = x_font

    return ws

def process_uploaded_file(uploaded_file):
    col_1, col_2, col_3 = st.columns([1, 3, 1])
    with col_2:
        try:
            df = pd.read_excel(
                uploaded_file,
                sheet_name='Sheet1',
                header=None,
                skiprows=15)
            
            st.subheader('')
            st.markdown("---")
            st.success(f"Successfully read data from **{uploaded_file.name}**.")

            # Filter the DataFrame using the 'Total Rooms' stop condition
            stop_row_index = df[df.iloc[:, 3].astype(str).str.contains('Total Rooms', na=False)].index
            if not stop_row_index.empty:
                df = df.iloc[:stop_row_index[0]]

            df = df.iloc[:, [3, 6, 7, 9]].copy()
            df.columns = ['Room_Raw', 'Guest_Name', 'Arrival_Date_Raw', 'Depart_Date_Raw']

            # Clean and process the columns
            # 1. Fill any NaN/blank values in 'Room_Raw' with a placeholder string ('0-')
            df['Room_Raw'] = df['Room_Raw'].fillna('0-').astype(str)

            # 2. Extract the room number (before the dash)
            df['Room_Number'] = df['Room_Raw'].str.split('-').str[0]

            # 3. Convert to integer (this is now safe because we filled the blanks with '0')
            df['Room_Number'] = df['Room_Number'].astype(int)

            # 4. Remove any placeholder rows added in step 1 (where Room_Number is 0)
            df.drop(df[df['Room_Number'] == 0].index, inplace=True)

            # Extract dates and format 
            df['Arrival_Date'] = pd.to_datetime(df['Arrival_Date_Raw'], format='%m/%d/%y', errors='coerce')
            df['Departure_Date'] = pd.to_datetime(df['Depart_Date_Raw'], format='%m/%d/%y', errors='coerce')
            df.dropna(subset=['Room_Number', 'Arrival_Date', 'Departure_Date'], inplace=True)

            # Rename Room_Number to 'Room'
            df.rename(columns={'Room_Number': 'Room'}, inplace=True)

            df.sort_values(by='Room', inplace=True)

            # Reorder columns if needed
            df = df[['Room', 'Guest_Name', 'Arrival_Date', 'Departure_Date']]

            # Reset index to start at 1 and name it 'CleanDex'
            df.index = range(1, len(df) + 1)
            df.index.name = 'CleanDex'

            # Build dictionary using 'Room' as key
            guest_data_dict = df.set_index('Room')[['Guest_Name', 'Arrival_Date', 'Departure_Date']].to_dict('index')

            df['Arrival_Date'] = df['Arrival_Date'].dt.strftime('%m/%d/%Y')
            df['Departure_Date'] = df['Departure_Date'].dt.strftime('%m/%d/%Y')

            # Step 6: Feedback
            st.info(f"Loaded **{len(guest_data_dict)}** guest records.")

            st.markdown("---")

            today = datetime.today()

            housekeeping_today_dict = {
                room: data
                for room, data in guest_data_dict.items()
                if should_get_housekeeping_today(data['Arrival_Date'], data['Departure_Date'], today)
            }

            st.subheader('')
            st.subheader("✅ Housekeeping Needed Today")
            if housekeeping_today_dict:
                st.markdown(f"🛏️ Rooms needing housekeeping: <span style='font-size:24px; font-weight:bold;'>{len(housekeeping_today_dict)}</span>",
                            unsafe_allow_html=True)
                df_today = df[df['Room'].isin(housekeeping_today_dict.keys())].copy()
                df_today.index = range(1, len(df_today) + 1)
                df_today.index.name = 'CleanDex'

                st.dataframe(df_today)

            else:
                st.success("🎉 No rooms need housekeeping today!")

            # Part 3 & 4: CREATE AND POPULATE NEW WORKBOOK
            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.page_setup.orientation = new_ws.ORIENTATION_LANDSCAPE
            new_ws = apply_excel_formatting(new_ws, guest_data_dict, today)

            # Part 5: PREPARE AND SAVE THE FINAL WORKBOOK for download
            output = BytesIO()
            new_wb.save(output)
            processed_data = output.getvalue()
            final_file_name = f"In House Guest List {date.today().strftime('%Y%m%d')}.xlsx"

            if st.download_button(
                label="Download Final Guest List (.xlsx)",
                data=processed_data,
                file_name=final_file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Click to download the newly created and formatted guest list file."):
                st.markdown(
                    "<div style='background-color:#e0f7fa;padding:15px;border-radius:10px;'>"
                    "<h3 style='color:#00796b;'>💾 File downloaded, Let's hit the sheets!</h3>"
                    "<p style='font-size:24px;color:#00332e;'>The spreadsheets! Jeez! 📋</p>"
                    "</div>",
                    unsafe_allow_html=True)
            else:
                st.markdown(
                    "<div style='background-color:#fff3e0;padding:15px;border-radius:10px;'>"
                    "<h4 style='color:#e65100;'>🫧 Sounds like we're getting into BUBBLE today!</h4>"
                    "<p style='font-size:16px;color:#3b1e00;'>Hit download to get a printable list. ✨</p>"
                    "</div>",
                    unsafe_allow_html=True)

        except Exception as e:
                st.error(f"An unexpected error occurred during processing: {e}")

st.set_page_config(layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
.center { display: flex; justify-content: center; text-align: center; }
</style>
""", unsafe_allow_html=True)

st.markdown("<h2 class='center' style='color:rgb(70, 130, 255);'>An EsteStyle Streamlit Page<br>Where Python Wiz Meets Data Biz!</h2>", unsafe_allow_html=True)
st.markdown("<img src='https://1drv.ms/i/s!ArWyPNkF5S-foZspwsary83MhqEWiA?embed=1&width=307&height=307' width='300' style='display: block; margin: 0 auto;'>", unsafe_allow_html=True)
st.markdown("<h3 class='center' style='color: rgb(135, 206, 250);'>🏨 Originally created for Best Western at Firestone 🛎️</h3>", unsafe_allow_html=True)
st.markdown("<h3 class='center' style='color: rgb(135, 206, 250);'>🤖 By Esteban C Loetz 📟</h3>", unsafe_allow_html=True)
st.markdown("##")

st.markdown("---")

st.markdown("<h2 class='center' style='color: rgb(112, 128, 140);'>🧼 Third Night's the Charm Automator 📋</h2>", unsafe_allow_html=True)
st.markdown("<h4 class='center'>This eliminates the need for arduous date mathing!</h4>", unsafe_allow_html=True)

col_1, col_2, col_3, col_4 = st.columns([.25, 3, 3, .25], gap="large")

with col_2:
    
    st.write("")
    st.markdown("""
    ### 🪜 Steps to Export Guest List:
    1. Open the '**Front Office**' user tab
    2. Select '**Reports**' from the top navigation bar
    3. Click '**Front Office**' tab that appears
    4. Hover over '**Reports**' (Bar Graph Icon) at right of screen
    5. Select '**In House Guest**' from dropdown menu
    6. Click '**Refresh**' button
    7. Click '**Export**' button
    8. Select '**Excel**'
    
    ##### 📁 An '**In_House_Guests**' file will be created in your Downloads folder.
    ##### 🔄 Use this file as input for file analysis section to the right.""")

# --- MAIN APP LOGIC ---
with col_3:
    st.write('')
    st.subheader("📥 Download file for analysis:")
    st.write('')
    st.markdown("Upload your **'In_House_Guests.xls'** file to generate a modified guest list with rooms needing service marked.")
    st.write('')
    uploaded_file = st.file_uploader(label="Upload guest list Excel file", type=['xls', 'xlsx'], label_visibility="collapsed")

if uploaded_file is not None:
    process_uploaded_file(uploaded_file)
else:
    st.info("Awaiting file upload... Once uploaded, your new file will be ready for download below.")
